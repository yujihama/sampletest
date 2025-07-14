"""
調書（エクセル形式）のレイアウト認識と記入箇所特定モジュール
"""

import os
import json
import base64
import logging
import tempfile
from pathlib import Path
from typing import Dict, List, Literal, Optional, TypedDict

# LangChain関連のインポート
from langchain_core.messages import HumanMessage
from langchain_openai import ChatOpenAI, AzureChatOpenAI
from pydantic import BaseModel, Field

# Excel操作関連のインポート
import openpyxl
from openpyxl.styles import PatternFill
import subprocess
import shutil  # soffice存在確認

from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Pydanticモデル: 調書の入力欄情報


class ExcelField(BaseModel):
    """Excelの入力欄情報を表すモデル"""
    cell_id: str = Field(..., description="セル番号（例: A1, B2）")
    key_name: str = Field(..., description="そのセルが表す情報の安定的なキー名（例: sample_id, test_result, note）。命名規則はスネークケース（例: sample_id）に従ってください。")
    description: str = Field(..., description="そのセルに記入すべき内容の説明")
    field_type: Literal["common", "sample"] = Field(
        ..., description="記入欄の種類（common: 手続き共通, sample: サンプル個別）")


class ExcelFormFields(BaseModel):
    """Excelフォームの入力欄情報のコレクション"""
    fields: List[ExcelField] = Field(..., description="検出された入力欄のリスト")
    reason: str = Field(..., description="判断根拠")


class ValidationResult(BaseModel):
    """入力欄の検証結果を表すモデル"""
    status: Literal["OK", "修正が必要"] = Field(..., description="検証結果のステータス")
    issues: Optional[List[str]] = Field(
        None, description="問題点のリスト（ステータスが「修正が必要」の場合）")
    suggestions: Optional[List[str]] = Field(
        None, description="修正提案のリスト（ステータスが「修正が必要」の場合）")


class FieldCorrection(BaseModel):
    """入力欄の修正指示"""
    add_fields: List[ExcelField] = Field(..., description="追加する入力欄のリスト")
    delete_fields: List[ExcelField] = Field(..., description="削除する入力欄のリスト")
    reason: str = Field(..., description="修正の判断根拠")


class ExcelFormatAnalyzer:
    """調書レイアウト認識クラス"""

    def __init__(
        self,
        model_name: str,
        api_key: str,
        max_iterations: int = 3,
        *,
        azure_api_key: str | None = None,
        azure_endpoint: str | None = None,
        azure_deployment: str | None = None,
        azure_api_version: str = "2023-07-01-preview",
    ):
        self.max_iterations = max_iterations
        if azure_endpoint and azure_api_key:
            self.llm = AzureChatOpenAI(
                api_key=azure_api_key,
                azure_endpoint=azure_endpoint,
                openai_api_version=azure_api_version,
                azure_deployment=azure_deployment or model_name,
            )
        else:
            self.llm = ChatOpenAI(
                model=model_name,
                api_key=api_key,
            )

    def analyze_format(self, excel_file_path: str, output_dir: str = None) -> Dict:
        """
        調書のレイアウトを認識し、記入箇所を特定する（反復的プロセス）

        Args:
            excel_file_path: 調書のExcelファイルパス
            output_dir: 出力ディレクトリ（指定しない場合はExcelファイルと同じディレクトリ）

        Returns:
            Dict: 分析結果
        """
        logger.info(f"調書レイアウト認識開始（反復的プロセス）: {excel_file_path}")

        try:
            # 出力ディレクトリの設定
            if output_dir:
                base_save_path = Path(output_dir)
            else:
                base_save_path = Path(excel_file_path).parent

            format_data_dir = base_save_path / "format_data"
            if format_data_dir.exists():
                import shutil
                shutil.rmtree(format_data_dir)
            format_data_dir.mkdir(exist_ok=True, parents=True)

            # 1. Excelデータの抽出とキャプチャ
            extracted_text, original_capture = self._extract_excel_data_and_capture(
                excel_file_path, format_data_dir
            )

            # "XXXX" と記載されたセルの検出
            placeholder_fields = self._detect_xxxx_fields(excel_file_path)

            if len(placeholder_fields) >= 3:
                # --- 簡易処理 ---
                structured_fields = self._estimate_fields_with_llm(
                    extracted_text, original_capture, 1
                )

                existing = {f.cell_id for f in structured_fields.fields}
                for pf in placeholder_fields:
                    if pf.cell_id not in existing:
                        structured_fields.fields.append(pf)

                highlighted_excel = self._highlight_fields(
                    excel_file_path, structured_fields, format_data_dir, 1
                )
                highlighted_captures = self._capture_highlighted_excel(
                    highlighted_excel, format_data_dir, 1
                )

                result = self._save_final_result(
                    structured_fields, format_data_dir, 1
                )

                return {
                    "status": "success",
                    "iterations": 1,
                    "final_validation_status": "N/A",
                    "fields": result["fields"],
                    "common_fields": result["common_fields"],
                    "sample_fields": result["sample_fields"],
                    "format_json_path": result["format_json_path"],
                    "highlighted_captures": highlighted_captures,
                }

            # --- 既存の反復処理 ---
            current_iteration = 1
            structured_fields = None
            validation_status = "修正が必要"
            validation_result = None

            while current_iteration <= self.max_iterations and validation_status != "OK":
                logger.info(
                    f"=== 反復 {current_iteration}/{self.max_iterations} 開始 ==="
                )

                if current_iteration == 1:
                    structured_fields = self._estimate_fields_with_llm(
                        extracted_text, original_capture, current_iteration
                    )
                else:
                    structured_fields = self._correct_fields_with_llm(
                        extracted_text,
                        original_capture,
                        structured_fields,
                        validation_result,
                        current_iteration,
                    )

                # 検出された"XXXX"セルを強制的に追加
                existing = {f.cell_id for f in structured_fields.fields}
                for pf in placeholder_fields:
                    if pf.cell_id not in existing:
                        structured_fields.fields.append(pf)

                highlighted_excel = self._highlight_fields(
                    excel_file_path, structured_fields, format_data_dir, current_iteration
                )

                highlighted_captures = self._capture_highlighted_excel(
                    highlighted_excel, format_data_dir, current_iteration
                )

                validation_result, validation_status = self._validate_with_llm(
                    original_capture, highlighted_captures[0], current_iteration
                )

                logger.info(
                    f"反復 {current_iteration} 検証結果: {validation_status}"
                )

                self._save_iteration_result(
                    structured_fields, validation_result, format_data_dir, current_iteration
                )

                if validation_status == "OK":
                    logger.info(f"検証成功！反復 {current_iteration} で完了")
                    break

                current_iteration += 1

            if validation_status != "OK":
                logger.warning(
                    f"最大反復回数 {self.max_iterations} に達しました。最後の結果を使用します。"
                )

            result = self._save_final_result(
                structured_fields, format_data_dir, current_iteration - 1
            )

            return {
                "status": "success",
                "iterations": current_iteration - 1,
                "final_validation_status": validation_status,
                "fields": result["fields"],
                "common_fields": result["common_fields"],
                "sample_fields": result["sample_fields"],
                "format_json_path": result["format_json_path"],
                "highlighted_captures": highlighted_captures,
            }

        except Exception as e:
            logger.error(f"調書レイアウト認識エラー: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return {
                "status": "error",
                "error_message": str(e)
            }

    def _extract_excel_data_and_capture(self, excel_file_path: str, output_dir: Path) -> tuple:
        """Excelデータの抽出とキャプチャ取得"""
        logger.info("Excelデータ抽出とキャプチャ開始")

        # キャプチャ用ディレクトリ作成
        captures_dir = output_dir / "captures"
        captures_dir.mkdir(exist_ok=True, parents=True)

        # Excelファイルを開く
        workbook = openpyxl.load_workbook(excel_file_path)

        # 抽出結果を格納するテキスト
        extracted_text = ""

        # 各シートの処理
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # シート名の追加
            extracted_text += f"## シート名: {sheet_name}\\n"

            # 結合セル情報の抽出
            merged_cells = []
            for merged_cell_range in sheet.merged_cells.ranges:
                merged_cells.append(str(merged_cell_range))

            if merged_cells:
                extracted_text += "### 結合セル情報:\\n"
                for cell_range in merged_cells:
                    extracted_text += f"- {cell_range}\\n"

            # セルデータの抽出
            extracted_text += "### セルデータ:\\n"
            extracted_text += "| セル | 値 | 書式 |\\n"
            extracted_text += "|-----|----|--------|\\n"

            for row in sheet.iter_rows():
                for cell in row:
                    # セルが空でない場合のみ処理
                    if cell.value is not None:
                        cell_addr = f"{cell.column_letter}{cell.row}"
                        cell_value = str(cell.value)

                        # 書式情報の取得
                        format_info = []
                        if cell.font.bold:
                            format_info.append("太字")
                        if cell.fill.fill_type == "solid":
                            fill_color = cell.fill.start_color.index
                            if fill_color != "00000000":  # デフォルト色でない場合
                                format_info.append(f"背景色:{fill_color}")

                        format_str = ", ".join(
                            format_info) if format_info else "-"

                        # テーブルに行を追加
                        extracted_text += f"| {cell_addr} | {cell_value} | {format_str} |\\n"

        # 抽出結果をファイルに保存
        extracted_text_file = output_dir / "extracted_excel_text.md"
        with open(extracted_text_file, "w", encoding="utf-8") as f:
            f.write(extracted_text)

        # Excelのキャプチャ取得
        original_capture = self._capture_excel(
            excel_file_path, captures_dir, "original_excel.png")

        logger.info(f"Excelデータ抽出完了: {extracted_text_file}")
        return extracted_text, original_capture

    def _capture_excel(self, excel_file_path: str, captures_dir: Path, output_name: str) -> str:
        """ExcelファイルのPNGキャプチャを取得"""
        # LibreOffice(soffice)存在確認
        if shutil.which("soffice") is None:
            logger.warning(
                "LibreOffice (soffice) が見つかりません。Excel のキャプチャをスキップします。")
            return ""

        try:
            # LibreOfficeを使用してPNGに変換
            command = f'soffice --headless --convert-to png "{excel_file_path}" --outdir "{captures_dir}"'
            subprocess.run(command, shell=True, check=True)

            # 生成されたPNGファイルを探す
            excel_basename = Path(excel_file_path).stem
            generated_files = list(captures_dir.glob(f"{excel_basename}*.png"))

            if generated_files:
                # 最初に見つかったファイルを使用
                generated_file = generated_files[0]
                output_path = captures_dir / output_name

                # ファイル名を変更
                if generated_file != output_path:
                    if output_path.exists():
                        output_path.unlink()
                    generated_file.rename(output_path)

                logger.info(f"Excelキャプチャ完了: {output_path}")
                return str(output_path)
            else:
                logger.error(f"キャプチャファイルが生成されませんでした: {excel_file_path}")
            return ""

        except Exception as e:
            logger.error(f"Excelキャプチャエラー: {str(e)}")
            return ""

    def _detect_xxxx_fields(self, excel_file_path: str) -> List[ExcelField]:
        """セルに 'XXXX' と記載されている箇所を入力欄として抽出"""
        workbook = openpyxl.load_workbook(excel_file_path)
        detected: List[ExcelField] = []
        counter = 1
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and "XXXX" in value:
                        # 左隣のセルから説明を推測
                        desc = None
                        if cell.col_idx > 1:
                            left = sheet.cell(row=cell.row, column=cell.col_idx - 1).value
                            if left and isinstance(left, str):
                                desc = f"{left} を記入する欄"
                        if not desc:
                            desc = "入力欄"
                        detected.append(
                            ExcelField(
                                cell_id=f"{cell.column_letter}{cell.row}",
                                key_name=f"placeholder_{counter}",
                                description=desc,
                                field_type="common",
                            )
                        )
                        counter += 1
        return detected

    def _estimate_fields_with_llm(self, extracted_text: str, image_path: str, iteration: int) -> ExcelFormFields:
        """LLMによる入力欄の推定（初回）"""
        logger.info(f"LLMによる入力欄推定開始（反復 {iteration}）")

        # 画像(base64)の準備 – キャプチャが失敗している場合は None
        base64_image: Optional[str] = None
        if image_path and Path(image_path).exists():
            try:
                with open(image_path, "rb") as img_file:
                    base64_image = base64.b64encode(
                        img_file.read()).decode("utf-8")
            except Exception as img_err:
                logger.warning(f"画像読み込みに失敗しました: {img_err}")
        else:
            logger.warning("キャプチャ画像が存在しないため、画像なしでLLMに問い合わせます")

        # structured_outputを使用
        llm_with_structure = self.llm.with_structured_output(ExcelFormFields, method="function_calling")

        # プロンプトの作成
        prompt = f"""
あなたは内部監査の調書レイアウト認識の専門家です。

以下はExcel調書から抽出したテキスト情報と、その調書の画像です。
この調書は内部監査で使用されるもので、以下の2種類の入力欄があります：

1. **手続き共通の記入欄**: 監査手続き名、実施日、実施者、監査対象などの共通情報
2. **各サンプルごとの記入欄**: サンプル番号、手続き結果、判定根拠、指摘事項などの個別情報

テキスト情報:
{extracted_text}

入力欄の特徴：
- 空白セル
- ラベル（太字や背景色付きのセル）の隣や下にある空白セル
- 表形式の場合、ヘッダー行の下の空白セル
- 既に値が入力されているセルでも、それが例や初期値と思われる場合は入力欄として扱う

各入力欄について、以下の情報を特定してください：
- `cell_id`: セル番号 (例: A1)
- `key_name`: そのセルが表す情報の安定的でユニークなキー名。命名規則はスネークケース（例: `sample_id`, `test_result`, `auditor_name`）に従ってください。これは後続の処理で機械的に利用します。
- `description`: 人間が理解しやすい、そのセルに記入すべき内容の説明。
- `field_type`: `common` (手続き共通) または `sample` (サンプル個別) のどちらか。

画像とテキスト情報の両方を参考にして、すべての入力欄を特定してください。
"""

        # LLMメッセージ構築
        message_content = [{"type": "text", "text": prompt}]
        if base64_image:
            message_content.append({
                "type": "image_url",
                "image_url": {"url": f"data:image/png;base64,{base64_image}"}
            })

        # マルチモーダルLLMに問い合わせ
        response = llm_with_structure.invoke([
            HumanMessage(content=message_content),
        ])

        logger.info(
            f"LLMによる入力欄推定完了（反復 {iteration}）: {len(response.fields)}個の入力欄を検出")
        return response

    def _highlight_fields(self, excel_file_path: str, structured_fields: ExcelFormFields,
                          output_dir: Path, iteration: int) -> str:
        """推定された入力欄をハイライト"""
        logger.info(f"入力欄のハイライト開始（反復 {iteration}）")

        # 元のExcelファイルをコピー
        workbook = openpyxl.load_workbook(excel_file_path)

        # ハイライト用フィル（共通欄: 黄色、サンプル欄: 水色）
        common_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色
        sample_fill = PatternFill(
            start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # 水色

        # 推定された入力欄をハイライト
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            for field in structured_fields.fields:
                try:
                    cell = sheet[field.cell_id]
                    original_value = cell.value

                    # field_typeに応じてハイライト色を変更
                    if field.field_type == "common":
                        cell.fill = common_fill
                        prefix = "共通"
                    else:
                        cell.fill = sample_fill
                        prefix = "サンプル"

                    # セルに識別情報を追加
                    if original_value is not None and str(original_value).strip() != "":
                        cell.value = f"{prefix}:{field.cell_id}:{original_value}"
                    else:
                        cell.value = f"{prefix}:{field.cell_id}"

                except Exception as cell_error:
                    logger.warning(
                        f"セル {field.cell_id} のハイライト中にエラー: {str(cell_error)}")

        # ハイライト済みExcelを保存
        highlighted_excel = output_dir / f"highlighted_excel_v{iteration}.xlsx"
        workbook.save(highlighted_excel)

        logger.info(f"入力欄のハイライト完了（反復 {iteration}）: {highlighted_excel}")
        return str(highlighted_excel)

    def _capture_highlighted_excel(self, highlighted_excel_path: str, output_dir: Path, iteration: int) -> List[str]:
        """ハイライト済みExcelのキャプチャ取得"""
        logger.info(f"ハイライト済みExcelキャプチャ開始（反復 {iteration}）")

        captures_dir = output_dir / "captures"
        captures_dir.mkdir(exist_ok=True, parents=True)

        # ハイライト済みExcelのキャプチャ
        highlighted_capture = self._capture_excel(
            highlighted_excel_path, captures_dir, f"highlighted_excel_v{iteration}.png"
        )

        return [highlighted_capture] if highlighted_capture else []

    def _validate_with_llm(self, original_capture: str, highlighted_capture: str, iteration: int) -> tuple:
        """LLMによる検証"""
        logger.info(f"LLMによる検証開始（反復 {iteration}）")

        # 画像をbase64エンコード
        with open(original_capture, "rb") as img_file:
            base64_original = base64.b64encode(img_file.read()).decode("utf-8")

        with open(highlighted_capture, "rb") as img_file:
            base64_highlighted = base64.b64encode(
                img_file.read()).decode("utf-8")

        # structured_outputを使用
        llm_with_structure = self.llm.with_structured_output(ValidationResult, method="function_calling")

        # プロンプトの作成
        prompt = f"""
以下は、Excel調書の画像と、入力欄として推定されたセルをハイライトした画像です。

ハイライトの色分け：
- 黄色: 手続き共通の記入欄（手続き名、実施日、実施者など）
- 水色: サンプル個別の記入欄（サンプル結果、判定根拠など）

このハイライトされた箇所について、以下の観点で評価を行ってください：
- 入力欄として適切なセルがハイライトされているか
- 入力すべきでない欄（ラベルや固定値）がハイライトされていないか
- 手続き共通欄とサンプル個別欄の分類が適切か
- 重要な入力欄が見落とされていないか

問題がなければステータスを「OK」としてください。
問題がある場合は、ステータスを「修正が必要」とし、具体的な問題点と修正案を説明してください。
"""

        # マルチモーダルLLMに問い合わせ
        response = llm_with_structure.invoke([
            HumanMessage(content=[
                {"type": "text", "text": prompt},
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/png;base64,{base64_original}"
                    }
                },
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/png;base64,{base64_highlighted}"
                    }
                }
            ])
        ])

        logger.info(f"LLMによる検証完了（反復 {iteration}）: {response.status} / {response.issues}")
        return response, response.status

    def _correct_fields_with_llm(self, extracted_text: str, original_capture: str,
                                 current_fields: ExcelFormFields, validation_result: ValidationResult,
                                 iteration: int) -> ExcelFormFields:
        """LLMによる入力欄の修正"""
        logger.info(f"LLMによる入力欄修正開始（反復 {iteration}）")

        # 画像をbase64エンコード
        with open(original_capture, "rb") as img_file:
            base64_image = base64.b64encode(img_file.read()).decode("utf-8")

        # structured_outputを使用
        llm_with_structure = self.llm.with_structured_output(FieldCorrection, method="function_calling")

        # プロンプトの作成
        prompt = f"""
あなたは内部監査の調書レイアウト認識の専門家です。

現在推定されている入力欄情報:
{current_fields.model_dump_json(indent=2)}

検証結果:
{validation_result.model_dump_json(indent=2)}

検証結果に基づいて、修正すべき箇所を特定してください：
- 追加すべき入力欄
- 削除すべき入力欄

修正指示を明確に回答してください。
追加するフィールドには、`cell_id`, `key_name`, `description`, `field_type` を含めてください。`key_name` はスネークケース（例: `sample_id`）で命名してください。
"""

        # マルチモーダルLLMに問い合わせ
        correction = llm_with_structure.invoke([
            HumanMessage(content=[
                {"type": "text", "text": prompt},
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/png;base64,{base64_image}"
                    }
                }
            ])
        ])

        # 現在のフィールドリストを取得
        current_fields_dict = {
            field.cell_id: field for field in current_fields.fields}

        # 削除するフィールドを処理
        for field_to_delete in correction.delete_fields:
            if field_to_delete.cell_id in current_fields_dict:
                del current_fields_dict[field_to_delete.cell_id]

        # 追加するフィールドを処理
        for field_to_add in correction.add_fields:
            current_fields_dict[field_to_add.cell_id] = field_to_add

        # 更新されたフィールドリストを作成
        updated_fields = list(current_fields_dict.values())

        # 更新されたExcelFormFieldsを作成
        updated_structured_fields = ExcelFormFields(
            fields=updated_fields,
            reason=correction.reason
        )

        logger.info(
            f"LLMによる入力欄修正完了（反復 {iteration}）: {len(updated_fields)}個の入力欄")
        return updated_structured_fields

    def _save_iteration_result(self, structured_fields: ExcelFormFields, validation_result: ValidationResult,
                               output_dir: Path, iteration: int):
        """反復結果の保存"""
        # 反復結果をファイルに保存
        iteration_result = {
            "iteration": iteration,
            "fields": [field.model_dump() for field in structured_fields.fields],
            "validation": validation_result.model_dump()
        }

        iteration_file = output_dir / f"iteration_{iteration}_result.json"
        with open(iteration_file, "w", encoding="utf-8") as f:
            json.dump(iteration_result, f, ensure_ascii=False, indent=2)

        logger.info(f"反復 {iteration} 結果保存: {iteration_file}")

    def _save_final_result(self, structured_fields: ExcelFormFields, output_dir: Path, final_iteration: int) -> Dict:
        """最終的な分析結果を保存"""
        logger.info("最終分析結果の保存")

        # commonとsampleに分類
        common_fields = {
            f.cell_id: {"description": f.description, "key_name": f.key_name}
            for f in structured_fields.fields if f.field_type == "common"
        }

        sample_fields_list = [
            f for f in structured_fields.fields if f.field_type == "sample"
        ]

        # sample_fieldsを構造化データに変換
        structured_sample_fields = {}
        if sample_fields_list:
            # セルを行と列に分解してソート
            sorted_sample_fields = sorted(
                sample_fields_list,
                key=lambda f: (
                    int("".join(filter(str.isdigit, f.cell_id))),
                    "".join(filter(str.isalpha, f.cell_id))
                )
            )
            
            start_row = int("".join(filter(str.isdigit, sorted_sample_fields[0].cell_id)))
            
            columns = {}
            processed_cols = set()
            for f in sorted_sample_fields:
                col_letter = "".join(filter(str.isalpha, f.cell_id))
                if col_letter not in processed_cols:
                    # 安定的なkey_nameをキーとして使用する
                    field_key = f.key_name
                    columns[field_key] = {
                        "col": col_letter,
                        "description": f.description
                    }
                    processed_cols.add(col_letter)

            structured_sample_fields = {
                "start_row": start_row,
                "columns": columns,
                 # 元のフィールドリストも保持しておく
                "fields": {f.cell_id: {"description": f.description, "key_name": f.key_name} for f in sample_fields_list}
            }


        final_result = {
            "fields": [f.dict() for f in structured_fields.fields],
            "common_fields": common_fields,
            "sample_fields": structured_sample_fields, # 構造化されたデータを格納
        }

        # JSONファイルに保存
        json_path = output_dir / "final_format_analysis_result.json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(final_result, f, ensure_ascii=False, indent=4)

        logger.info(f"最終結果を保存しました: {json_path}")
        
        # ハイライト画像を最終版としてコピー
        final_capture_path = None
        src_capture = output_dir / "captures" / f"highlighted_excel_v{final_iteration}.png"
        if src_capture.exists():
            final_capture_path = output_dir / "captures" / "highlighted_excel_final.png"
            shutil.copy(src_capture, final_capture_path)


        # 返り値にパスを追加
        final_result["format_json_path"] = str(json_path)
        final_result["highlighted_capture_path"] = str(final_capture_path) if final_capture_path else None

        return final_result
